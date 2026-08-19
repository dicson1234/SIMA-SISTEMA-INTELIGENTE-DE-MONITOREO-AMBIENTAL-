"""
SIMA — Sistema Inteligente de Monitoreo Ambiental
Módulo de Gestión de Archivos Excel

Utiliza openpyxl para crear, actualizar y dar formato profesional a archivos
de reporte Excel (.xlsx). Cada archivo generado contiene una pestaña de resumen
estadístico y otra de datos crudos ordenados cronológicamente.

Responsabilidades:
    - Crear el archivo Excel estructurado a partir de los datos recolectados.
    - Aplicar estilos profesionales (fuentes, rellenos, bordes, alineación).
    - Ajustar automáticamente el ancho de las columnas.
    - Crear hojas de datos ("Datos de Monitoreo") y de estadísticas ("Resumen").
    - Guardar de forma segura en el directorio especificado.

Autor:  Equipo SIMA — Ingeniero de Software
Fecha:  2026-07-14
"""

from datetime import datetime
from pathlib import Path
from typing import List, Dict, Any, Union
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from config import EXCEL_DIR, EXCEL_COLUMNS, EXCEL_FILENAME_FORMAT, FILE_DATE_FORMAT, FILE_TIME_FORMAT
from sensor_manager import SensorReading
from logger_manager import get_logger, log_data_event

# Logger del módulo
logger = get_logger(__name__)


class ExcelManager:
    """Administra la creación y formateo de archivos de reporte Excel para SIMA."""

    def __init__(self, output_directory: Union[str, Path] = EXCEL_DIR) -> None:
        """Inicializa el gestor de Excel.

        Args:
            output_directory: Carpeta destino para guardar los archivos.
        """
        self.output_dir: Path = Path(output_directory)
        self.output_dir.mkdir(parents=True, exist_ok=True)

    def generate_report(
        self,
        readings: List[SensorReading],
        stats: Dict[str, Any]
    ) -> Path:
        """Genera un reporte Excel completo con datos históricos y resumen estadístico.

        Args:
            readings: Lista de objetos SensorReading con el historial de mediciones.
            stats: Diccionario con las estadísticas de la sesión (del StatisticsManager).

        Returns:
            La ruta absoluta al archivo generado.
        """
        if not readings:
            raise ValueError("No se pueden generar reportes de Excel sin lecturas cargadas.")

        # Generar nombre dinámico basado en la fecha y hora de la primera lectura
        first_time = readings[0].timestamp
        date_str = first_time.strftime(FILE_DATE_FORMAT)
        time_str = first_time.strftime(FILE_TIME_FORMAT)
        filename = EXCEL_FILENAME_FORMAT.format(date=date_str, time=time_str)
        file_path = self.output_dir / filename

        # Crear libro de trabajo
        wb = openpyxl.Workbook()
        
        # 1. Hoja de Resumen (Primera pestaña)
        ws_summary = wb.active
        ws_summary.title = "Resumen"
        ws_summary.views.sheetView[0].showGridLines = True
        self._populate_summary_sheet(ws_summary, stats, readings)

        # 2. Hoja de Datos (Segunda pestaña)
        ws_data = wb.create_sheet(title="Datos de Monitoreo")
        ws_data.views.sheetView[0].showGridLines = True
        self._populate_data_sheet(ws_data, readings)

        # Guardar en disco
        wb.save(file_path)
        log_data_event("EXCEL_GUARDADO", str(file_path))
        logger.info("Reporte Excel generado exitosamente: %s", file_path.name)
        
        return file_path

    def _populate_summary_sheet(
        self,
        ws: openpyxl.worksheet.worksheet.Worksheet,
        stats: Dict[str, Any],
        readings: List[SensorReading]
    ) -> None:
        """Llena la hoja de resumen estadístico con un formato elegante e industrial."""
        # Estilos y Paleta de Colores (Corporativo SIMA azul/gris)
        fill_header = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid") # Slate Gray oscuro
        fill_sub_header = PatternFill(start_color="374151", end_color="374151", fill_type="solid")
        fill_accent = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid") # Azul muy claro
        
        font_title = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
        font_section = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
        font_bold = Font(name="Calibri", size=11, bold=True)
        font_regular = Font(name="Calibri", size=11)
        
        align_center = Alignment(horizontal="center", vertical="center")
        align_left = Alignment(horizontal="left", vertical="center")
        align_right = Alignment(horizontal="right", vertical="center")
        
        thin_border = Border(
            left=Side(style='thin', color='D1D5DB'),
            right=Side(style='thin', color='D1D5DB'),
            top=Side(style='thin', color='D1D5DB'),
            bottom=Side(style='thin', color='D1D5DB')
        )

        # Título principal de la hoja
        ws.merge_cells("A1:D2")
        ws["A1"] = "SIMA — RESUMEN ESTADÍSTICO DE MONITOREO"
        ws["A1"].font = font_title
        ws["A1"].alignment = align_center
        ws["A1"].fill = fill_header

        # Sección: Información General
        ws.merge_cells("A4:D4")
        ws["A4"] = "Información General del Reporte"
        ws["A4"].font = font_section
        ws["A4"].fill = fill_sub_header
        ws["A4"].alignment = align_left

        general_info = [
            ("Fecha de Inicio:", readings[0].timestamp_str),
            ("Fecha de Fin:", readings[-1].timestamp_str),
            ("Total de Muestras:", stats.get("sample_count", 0)),
            ("Tiempo de Monitoreo:", stats.get("elapsed_time_str", "00:00:00"))
        ]

        for i, (label, val) in enumerate(general_info, start=5):
            ws.cell(row=i, column=1, value=label).font = font_bold
            ws.cell(row=i, column=2, value=val).font = font_regular
            ws.cell(row=i, column=1).border = thin_border
            ws.cell(row=i, column=2).border = thin_border
            # Combinar filas adyacentes para alinearlo bien
            ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=4)
            for col in range(2, 5):
                ws.cell(row=i, column=col).border = thin_border

        # Sección: Métricas Estadísticas
        ws.merge_cells("A10:E10")
        ws["A10"] = "Estadísticas de Variables Ambientales"
        ws["A10"].font = font_section
        ws["A10"].fill = fill_sub_header
        ws["A10"].alignment = align_left

        # Encabezados de Tabla de Métricas
        headers = ["Métrica", "Temperatura", "Humedad", "Luminosidad", "Unidad"]
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=11, column=col_idx, value=header)
            cell.font = font_bold
            cell.fill = fill_accent
            cell.alignment = align_center
            cell.border = thin_border

        metric_rows = [
            ("Máximo", stats.get("temp_max"), stats.get("hum_max"), stats.get("light_max"), "°C / % / Lux"),
            ("Mínimo", stats.get("temp_min"), stats.get("hum_min"), stats.get("light_min"), "°C / % / Lux"),
            ("Promedio", stats.get("temp_avg"), stats.get("hum_avg"), stats.get("light_avg"), "°C / % / Lux")
        ]

        for row_idx, (metric, t_val, h_val, l_val, unit) in enumerate(metric_rows, start=12):
            c1 = ws.cell(row=row_idx, column=1, value=metric)
            c2 = ws.cell(row=row_idx, column=2, value=t_val)
            c3 = ws.cell(row=row_idx, column=3, value=h_val)
            c4 = ws.cell(row=row_idx, column=4, value=l_val)
            c5 = ws.cell(row=row_idx, column=5, value=unit)
            
            c1.font = font_bold
            c2.font = font_regular
            c3.font = font_regular
            c4.font = font_regular
            c5.font = font_regular

            c1.alignment = align_left
            c2.alignment = align_right
            c3.alignment = align_right
            c4.alignment = align_right
            c5.alignment = align_center

            for cell in (c1, c2, c3, c4, c5):
                cell.border = thin_border

        # Auto-ajuste de ancho de columnas
        self._autofit_columns(ws)

    def _populate_data_sheet(
        self,
        ws: openpyxl.worksheet.worksheet.Worksheet,
        readings: List[SensorReading]
    ) -> None:
        """Llena la hoja con la tabla de datos de monitoreo detallada."""
        # Estilos
        fill_header = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid") # Azul brillante
        font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        font_regular = Font(name="Calibri", size=11)
        align_center = Alignment(horizontal="center", vertical="center")
        align_right = Alignment(horizontal="right", vertical="center")
        
        thin_border = Border(
            left=Side(style='thin', color='E5E7EB'),
            right=Side(style='thin', color='E5E7EB'),
            top=Side(style='thin', color='E5E7EB'),
            bottom=Side(style='thin', color='E5E7EB')
        )

        # Cabecera de la tabla
        for col_idx, column_name in enumerate(EXCEL_COLUMNS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=column_name)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = align_center
            cell.border = thin_border

        # Insertar registros
        for row_idx, r in enumerate(readings, start=2):
            c1 = ws.cell(row=row_idx, column=1, value=r.timestamp_str)
            c2 = ws.cell(row=row_idx, column=2, value=r.temperature)
            c3 = ws.cell(row=row_idx, column=3, value=r.humidity)
            c4 = ws.cell(row=row_idx, column=4, value=r.light)
            c5 = ws.cell(row=row_idx, column=5, value=r.temp_class.label)
            c6 = ws.cell(row=row_idx, column=6, value=r.hum_class.label)
            c7 = ws.cell(row=row_idx, column=7, value=r.light_class.label)
            c8 = ws.cell(row=row_idx, column=8, value=r.comfort.score)

            # Formatos numéricos
            c2.number_format = '0.0'
            c3.number_format = '0.0'
            c4.number_format = '0.0'
            c8.number_format = '0.0'

            # Alineaciones
            c1.alignment = align_center
            c2.alignment = align_right
            c3.alignment = align_right
            c4.alignment = align_right
            c5.alignment = align_center
            c6.alignment = align_center
            c7.alignment = align_center
            c8.alignment = align_right

            # Aplicar fuentes y bordes
            for cell in (c1, c2, c3, c4, c5, c6, c7, c8):
                cell.font = font_regular
                cell.border = thin_border

            # Colores discretos según confort
            if r.comfort.score >= 80:
                c8.fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid") # Light Green
            elif r.comfort.score < 40:
                c8.fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid") # Light Red

        # Auto-ajuste de ancho de columnas
        self._autofit_columns(ws)

    @staticmethod
    def _autofit_columns(ws: openpyxl.worksheet.worksheet.Worksheet) -> None:
        """Ajusta automáticamente el ancho de las columnas según su contenido para evitar textos truncados."""
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                # Si la celda está combinada, ignorar el cálculo del tamaño de esta celda
                # para evitar deformar las columnas
                if cell.coordinate in ws.merged_cells:
                    continue
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            # Ajustar ancho con un margen de holgura
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
